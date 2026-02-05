# ui_modern_fullscreen.py
import json, threading, queue, serial, time, os, csv
from datetime import datetime
import customtkinter as ctk

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

PORT = "/dev/ttyUSB0"
BAUD = 115200
q = queue.Queue()

ser = None
ser_lock = threading.Lock()

# --------------------------
# Comms fault screen
# --------------------------
ENABLE_FAULT_SCREEN = False          # set False to disable overlay (debug)
RX_TIMEOUT_S = 3.0                  # show fault if no RX for this many seconds

connected = False
last_rx_time = 0.0
last_disconnect_reason = ""

# --------------------------
# Logging
# --------------------------
LOG_DIR = "logs"
os.makedirs(LOG_DIR, exist_ok=True)

def now_iso():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")[:-3]

def log_base_name():
    return f"swarm_log_{datetime.now().strftime('%Y-%m-%d')}"

XLSX_PATH = os.path.join(LOG_DIR, log_base_name() + ".xlsx")
CSV_PATH  = os.path.join(LOG_DIR, log_base_name() + ".csv")

EVENT_HEADERS = [
    "timestamp",
    "event_type",
    "raw",
    "C", "F", "flowLpm", "turbidity_pct", "NTU",
    "heaterEnable", "setTempF", "heaterDemand", "heaterOn",
    "overTemp", "secOver", "heaterStop",
    "levelOk", "lidClosed",
    "ui_heater_enable",
    "ui_cycle_state", "ui_cycle_remaining_s", "ui_cycle_duration_s", "ui_cycle_temp_set_f",
]

class AsyncLogger:
    def __init__(self, xlsx_path: str, csv_path: str):
        self.xlsx_path = xlsx_path
        self.csv_path = csv_path
        self.lq = queue.Queue()
        self.stop_flag = threading.Event()

        self._init_csv()
        self._init_xlsx()

        self.t = threading.Thread(target=self._worker, daemon=True)
        self.t.start()

    def _init_csv(self):
        if not os.path.exists(self.csv_path):
            with open(self.csv_path, "w", newline="", encoding="utf-8") as f:
                csv.writer(f).writerow(EVENT_HEADERS)

    def _init_xlsx(self):
        if os.path.exists(self.xlsx_path):
            return
        wb = Workbook()
        ws_events = wb.active
        ws_events.title = "Events"
        ws_events.append(EVENT_HEADERS)

        ws_rx = wb.create_sheet("RxRaw")
        ws_rx.append(["timestamp", "raw_line"])

        ws_tx = wb.create_sheet("TxCmd")
        ws_tx.append(["timestamp", "command"])

        for ws in (ws_events, ws_rx, ws_tx):
            for i, h in enumerate(ws[1], start=1):
                ws.column_dimensions[get_column_letter(i)].width = max(12, min(42, len(str(h)) + 2))

        wb.save(self.xlsx_path)

    def log_event(self, row: dict):
        self.lq.put(row)

    def _append_xlsx(self, batch_rows: list[dict]):
        wb = load_workbook(self.xlsx_path)
        ws_events = wb["Events"]
        ws_rx = wb["RxRaw"]
        ws_tx = wb["TxCmd"]

        for r in batch_rows:
            et = r.get("event_type", "")
            ts = r.get("timestamp", "")
            raw = r.get("raw", "")

            ws_events.append([r.get(k, "") for k in EVENT_HEADERS])

            if et == "rx_raw":
                ws_rx.append([ts, raw])
            elif et == "tx_cmd":
                ws_tx.append([ts, raw])

        wb.save(self.xlsx_path)

    def _append_csv(self, batch_rows: list[dict]):
        with open(self.csv_path, "a", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            for r in batch_rows:
                w.writerow([r.get(k, "") for k in EVENT_HEADERS])

    def _worker(self):
        batch = []
        last_flush = time.time()
        FLUSH_EVERY_SEC = 2.0
        FLUSH_EVERY_ROWS = 25

        while not self.stop_flag.is_set():
            try:
                batch.append(self.lq.get(timeout=0.25))
            except queue.Empty:
                pass

            do_flush = False
            if batch and len(batch) >= FLUSH_EVERY_ROWS:
                do_flush = True
            if batch and (time.time() - last_flush) >= FLUSH_EVERY_SEC:
                do_flush = True

            if do_flush:
                try:
                    self._append_csv(batch)
                    self._append_xlsx(batch)
                except Exception:
                    try:
                        self._append_csv(batch)
                    except Exception:
                        pass
                batch.clear()
                last_flush = time.time()

    def stop(self):
        self.stop_flag.set()

logger = AsyncLogger(XLSX_PATH, CSV_PATH)

def log_status(msg: str):
    logger.log_event({"timestamp": now_iso(), "event_type": "status", "raw": msg})

def log_rx(raw_line: str):
    logger.log_event({"timestamp": now_iso(), "event_type": "rx_raw", "raw": raw_line})

def log_tx(cmd: str):
    logger.log_event({"timestamp": now_iso(), "event_type": "tx_cmd", "raw": cmd})

# --------------------------
# Serial thread
# --------------------------
def reader():
    global ser
    while True:
        try:
            ser = serial.Serial(PORT, BAUD, timeout=1)
            q.put(("status", f"Connected: {PORT}"))
            log_status(f"Connected: {PORT}")

            global connected, last_rx_time, last_disconnect_reason
            connected = True
            last_disconnect_reason = ""
            last_rx_time = time.time()

            while True:
                line = ser.readline().decode("utf-8", "ignore").strip()
                if line:
                    last_rx_time = time.time()
                    q.put(("data", line))
                    log_rx(line)
        except Exception as e:
            connected = False
            last_disconnect_reason = str(e)
            q.put(("status", f"Disconnected: {e}"))
            log_status(f"Disconnected: {e}")
            try:
                if ser:
                    ser.close()
            except:
                pass
            ser = None
            time.sleep(1)

def send(cmd):
    global ser
    log_tx(cmd)
    try:
        with ser_lock:
            if ser and ser.is_open:
                ser.write((cmd + "\n").encode("utf-8"))
            else:
                q.put(("status", "Send failed: serial not connected"))
                log_status("Send failed: serial not connected")
    except Exception as e:
        q.put(("status", f"Send failed: {e}"))
        log_status(f"Send failed: {e}")

def main():
    ctk.set_appearance_mode("light")
    ctk.set_default_color_theme("blue")

    app = ctk.CTk()
    app.update_idletasks()
    app.attributes("-fullscreen", True)
    app.title("SWARM")

    def on_close(_=None):
        try:
            logger.stop()
        except:
            pass
        app.destroy()

    app.bind("<Escape>", on_close)

    # --------------------------
    # Top-level layout: left nav + main content
    # --------------------------
    root = ctk.CTkFrame(app, corner_radius=0, fg_color="white")
    root.pack(fill="both", expand=True)
    root.grid_rowconfigure(0, weight=1)
    root.grid_columnconfigure(1, weight=1)

    nav = ctk.CTkFrame(root, corner_radius=0, fg_color="#F6F7F9", width=220)
    nav.grid(row=0, column=0, sticky="nsw")
    nav.grid_propagate(False)

    main_area = ctk.CTkFrame(root, corner_radius=0, fg_color="white")
    main_area.grid(row=0, column=1, sticky="nsew")
    main_area.grid_rowconfigure(1, weight=1)
    main_area.grid_columnconfigure(0, weight=1)

    # Header inside main area
    header = ctk.CTkFrame(main_area, corner_radius=0, fg_color="white")
    header.grid(row=0, column=0, sticky="ew", padx=20, pady=(16, 8))
    header.grid_columnconfigure(1, weight=1)

    ctk.CTkLabel(header, text="Control Panel", text_color="black",
                 font=("SF Pro Display", 28, "bold")).grid(row=0, column=0, sticky="w")

    title_status = ctk.StringVar(value="")
    ctk.CTkLabel(header, textvariable=title_status, text_color="black",
                 font=("SF Pro Text", 14)).grid(row=0, column=2, sticky="e")

    page_container = ctk.CTkFrame(main_area, corner_radius=0, fg_color="white")
    page_container.grid(row=1, column=0, sticky="nsew", padx=20, pady=(0, 16))
    page_container.grid_rowconfigure(0, weight=1)
    page_container.grid_columnconfigure(0, weight=1)

    # --------------------------
    # Fault overlay (full-screen)
    # --------------------------
    fault_overlay = ctk.CTkFrame(root, corner_radius=0, fg_color="white")
    fault_overlay.grid(row=0, column=0, columnspan=2, sticky="nsew")
    fault_overlay.grid_remove()  # start hidden

    fault_overlay.grid_rowconfigure(0, weight=1)
    fault_overlay.grid_rowconfigure(2, weight=1)
    fault_overlay.grid_columnconfigure(0, weight=1)

    fault_title = ctk.StringVar(value="COMMUNICATION FAULT")
    fault_detail = ctk.StringVar(value="")

    ctk.CTkLabel(
        fault_overlay, textvariable=fault_title,
        font=("SF Pro Display", 40, "bold"), text_color="black"
    ).grid(row=1, column=0, sticky="s", pady=(0, 10))

    ctk.CTkLabel(
        fault_overlay, textvariable=fault_detail,
        font=("SF Pro Text", 18), text_color="gray30"
    ).grid(row=2, column=0, sticky="n", pady=(0, 30))

    def show_fault_overlay(reason: str):
        fault_detail.set(reason)
        fault_overlay.tkraise()
        fault_overlay.grid()

    def hide_fault_overlay():
        fault_overlay.grid_remove()

    def comms_watchdog():
        if not ENABLE_FAULT_SCREEN:
            hide_fault_overlay()
            app.after(250, comms_watchdog)
            return

        now = time.time()
        timed_out = (last_rx_time > 0.0) and ((now - last_rx_time) > RX_TIMEOUT_S)

        if (not connected) or timed_out:
            if not connected and last_disconnect_reason:
                reason = f"ESP32 disconnected, {last_disconnect_reason}"
            elif timed_out:
                reason = f"No telemetry received for {RX_TIMEOUT_S:.1f} seconds"
            else:
                reason = "Communication fault"
            show_fault_overlay(reason)
        else:
            hide_fault_overlay()

        app.after(250, comms_watchdog)


    # --------------------------
    # Operation Panel state
    # --------------------------
    # Machine state labels for operator (simple and explicit).
    # Map whatever your controller uses later.
    # IDLE | WASHING | PAUSED | COMPLETE | FAULT
    cycle_state = ctk.StringVar(value="IDLE")

    cycle_duration_min = ctk.IntVar(value=10)
    cycle_temp_set_f = ctk.DoubleVar(value=95.0)
    cycle_remaining_s = ctk.IntVar(value=0)

    def log_cycle_snapshot():
        logger.log_event({
            "timestamp": now_iso(),
            "event_type": "status",
            "raw": "UI cycle state update",
            "ui_cycle_state": cycle_state.get(),
            "ui_cycle_remaining_s": cycle_remaining_s.get(),
            "ui_cycle_duration_s": int(cycle_duration_min.get() * 60),
            "ui_cycle_temp_set_f": float(cycle_temp_set_f.get()),
        })

    _cycle_tick_job = None
    def _cycle_tick():
        nonlocal _cycle_tick_job
        if cycle_state.get() == "WASHING":
            r = cycle_remaining_s.get()
            if r > 0:
                cycle_remaining_s.set(r - 1)
            else:
                cycle_state.set("COMPLETE")
                log_cycle_snapshot()
        _cycle_tick_job = app.after(1000, _cycle_tick)

    def start_cycle():
        if cycle_state.get() in ("WASHING", "PAUSED"):
            return
        cycle_state.set("WASHING")
        cycle_remaining_s.set(int(cycle_duration_min.get() * 60))
        log_cycle_snapshot()
        # placeholder: send(f"CYCLE START {cycle_remaining_s.get()} {cycle_temp_set_f.get():.1f}")

    def pause_resume_cycle():
        if cycle_state.get() == "WASHING":
            cycle_state.set("PAUSED")
            log_cycle_snapshot()
            # placeholder: send("CYCLE PAUSE")
        elif cycle_state.get() == "PAUSED":
            cycle_state.set("WASHING")
            log_cycle_snapshot()
            # placeholder: send("CYCLE RESUME")

    def stop_cycle():
        if cycle_state.get() in ("IDLE", "COMPLETE"):
            return
        cycle_state.set("IDLE")
        cycle_remaining_s.set(0)
        log_cycle_snapshot()
        # placeholder: send("CYCLE STOP")

    _cycle_tick()

    def is_cycle_active() -> bool:
        return cycle_state.get() in ("WASHING", "PAUSED")

    # --------------------------
    # Pages
    # --------------------------
    page_operation = ctk.CTkFrame(page_container, corner_radius=0, fg_color="white")
    page_status    = ctk.CTkFrame(page_container, corner_radius=0, fg_color="white")

    for p in (page_operation, page_status):
        p.grid(row=0, column=0, sticky="nsew")

    def show_page(name: str):
        if name == "operation":
            page_operation.tkraise()
        else:
            page_status.tkraise()

    # --------------------------
    # Left navigation (static)
    # --------------------------
    ctk.CTkLabel(nav, text="SWARM", font=("SF Pro Text", 48, "bold"), text_color="black").pack(
        anchor="w", padx=16, pady=(18, 8)
    )

    # Container to vertically center nav buttons
    nav_btns = ctk.CTkFrame(nav, corner_radius=0, fg_color="transparent")
    nav_btns.pack(fill="both", expand=True, padx=0, pady=0)

    nav_btns.grid_rowconfigure(0, weight=1)  # top spacer
    nav_btns.grid_rowconfigure(3, weight=1)  # bottom spacer
    nav_btns.grid_columnconfigure(0, weight=1)

    def nav_button(parent, text, command):
        return ctk.CTkButton(
            parent, text=text, command=command,
            height=64, corner_radius=10,
            fg_color="#FFFFFF", hover_color="#EDEFF3",
            text_color="black", anchor="w"
        )

    # spacer
    ctk.CTkFrame(nav_btns, fg_color="transparent").grid(row=0, column=0, sticky="nsew")

    btn_op = nav_button(nav_btns, "Operation", lambda: show_page("operation"))
    btn_op.grid(row=1, column=0, sticky="ew", padx=12, pady=(0, 10))

    btn_status = nav_button(nav_btns, "Status", lambda: show_page("status"))
    btn_status.grid(row=2, column=0, sticky="ew", padx=12, pady=(0, 10))

    # spacer
    ctk.CTkFrame(nav_btns, fg_color="transparent").grid(row=3, column=0, sticky="nsew")

    # --------------------------
    # Operation Page
    # --------------------------
    # Operation page: 1 column, 3 equal-height rows
    page_operation.grid_columnconfigure(0, weight=1)
    page_operation.grid_rowconfigure(0, weight=1)
    page_operation.grid_rowconfigure(1, weight=1)
    page_operation.grid_rowconfigure(2, weight=1)

    # (1) Replace description card with simple Status card
    status_card = ctk.CTkFrame(page_operation, corner_radius=16, fg_color="#F6F7F9")
    status_card.grid(row=0, column=0, sticky="nsew", pady=(0, 12))
    status_card.grid_columnconfigure(0, weight=1)

    ctk.CTkLabel(status_card, text="Machine Status", font=("SF Pro Text", 16, "bold"), text_color="black").grid(
        row=0, column=0, sticky="w", padx=16, pady=(14, 6)
    )
    ctk.CTkLabel(status_card, textvariable=cycle_state, font=("SF Pro Display", 30, "bold"), text_color="black").grid(
        row=1, column=0, sticky="w", padx=16, pady=(0, 14)
    )

    # Row 1: Only show one card at a time
    set_card = ctk.CTkFrame(page_operation, corner_radius=16)
    time_card = ctk.CTkFrame(page_operation, corner_radius=16)

    # Setpoints card (shown only when IDLE/COMPLETE/FAULT)
    set_card.grid_columnconfigure(1, weight=1)
    ctk.CTkLabel(set_card, text="Setpoints", font=("SF Pro Text", 16, "bold")).grid(
        row=0, column=0, columnspan=2, sticky="w", padx=14, pady=(14, 10)
    )
    ctk.CTkLabel(set_card, text="Cycle Duration (min)", font=("SF Pro Text", 13), text_color="gray30").grid(
        row=1, column=0, sticky="w", padx=14, pady=(0, 6)
    )
    ctk.CTkEntry(set_card, textvariable=cycle_duration_min, width=120).grid(
        row=1, column=1, sticky="e", padx=14, pady=(0, 6)
    )

    ctk.CTkLabel(set_card, text="Temperature Setpoint (°F)", font=("SF Pro Text", 13), text_color="gray30").grid(
        row=2, column=0, sticky="w", padx=14, pady=(0, 14)
    )
    ctk.CTkEntry(set_card, textvariable=cycle_temp_set_f, width=120).grid(
        row=2, column=1, sticky="e", padx=14, pady=(0, 14)
    )

    # Time remaining card (shown only when WASHING/PAUSED)
    time_card.grid_columnconfigure(0, weight=1)
    ctk.CTkLabel(time_card, text="Time Remaining", font=("SF Pro Text", 16, "bold")).grid(
        row=0, column=0, sticky="w", padx=14, pady=(14, 10)
    )

    def fmt_mmss(s: int):
        m = max(0, int(s)) // 60
        sec = max(0, int(s)) % 60
        return f"{m:02d}:{sec:02d}"

    remaining_text = ctk.StringVar(value="00:00")

    ctk.CTkLabel(time_card, textvariable=remaining_text, font=("SF Pro Display", 34, "bold")).grid(
        row=1, column=0, sticky="w", padx=14, pady=(0, 14)
    )

    def update_remaining_label():
        remaining_text.set(fmt_mmss(cycle_remaining_s.get()))
        app.after(250, update_remaining_label)
    update_remaining_label()

    def update_op_row_cards():
        # (2) Only show one: set_card if idle-ish, time_card if active
        if is_cycle_active():
            set_card.grid_forget()
            time_card.grid(row=1, column=0, sticky="nsew", pady=(0, 12))
        else:
            time_card.grid_forget()
            set_card.grid(row=1, column=0, sticky="nsew", pady=(0, 12))
        app.after(250, update_op_row_cards)
    update_op_row_cards()

    # Actions
    action = ctk.CTkFrame(page_operation, corner_radius=16, fg_color="#F6F7F9")
    action.grid(row=2, column=0, sticky="nsew")
    action.grid_columnconfigure((0,1,2), weight=1)

    # Make the single row expand vertically
    action.grid_rowconfigure(0, weight=1)

    # Make columns expand; we will reconfigure weights depending on state
    action.grid_columnconfigure(0, weight=1)
    action.grid_columnconfigure(1, weight=1)
    action.grid_columnconfigure(2, weight=1)

    btn_start = ctk.CTkButton(action, text="Start", corner_radius=12, command=start_cycle)
    btn_pause = ctk.CTkButton(action, text="Pause / Resume", corner_radius=12, command=pause_resume_cycle)
    btn_stop  = ctk.CTkButton(action, text="Stop", corner_radius=12, command=stop_cycle)

    _last_action_mode = None  # "idle" or "active"
    def update_action_buttons():
        nonlocal _last_action_mode

        mode = "idle" if cycle_state.get() in ("IDLE", "COMPLETE") else "active"
        if mode == _last_action_mode:
            app.after(200, update_action_buttons)
            return

        _last_action_mode = mode

        # Hide all once (only when mode changes)
        btn_start.grid_forget()
        btn_pause.grid_forget()
        btn_stop.grid_forget()

        if mode == "idle":
            # Only Start, spanning full width
            action.grid_columnconfigure(0, weight=1)
            action.grid_columnconfigure(1, weight=0)
            action.grid_columnconfigure(2, weight=0)

            btn_start.grid(row=0, column=0, columnspan=3, padx=12, pady=12, sticky="nsew")
        else:
            # Washing or paused: Pause/Resume + Stop
            action.grid_columnconfigure(0, weight=0)
            action.grid_columnconfigure(1, weight=1)
            action.grid_columnconfigure(2, weight=1)

            btn_pause.grid(row=0, column=1, padx=12, pady=12, sticky="nsew")
            btn_stop.grid(row=0, column=2, padx=12, pady=12, sticky="nsew")

        app.after(200, update_action_buttons)

    update_action_buttons()

    # --------------------------
    # Status Page (telemetry) - motor card removed
    # --------------------------
    content = ctk.CTkFrame(page_status, corner_radius=0, fg_color="white")
    content.pack(fill="both", expand=True, padx=0, pady=0)
    content.grid_columnconfigure((0,1), weight=1)
    content.grid_rowconfigure((0,1), weight=1)

    # temp card
    tempCard = ctk.CTkFrame(content, corner_radius=16)
    tempCard.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)
    ctk.CTkLabel(tempCard, text="Temperature", font=("SF Pro Text",16,"bold")).grid(row=0, column=0, sticky="w", padx=12, pady=(12,4))
    tempNowF = ctk.StringVar(value="—")
    tempSetF = ctk.StringVar(value="Set: —")

    ctk.CTkLabel(tempCard, textvariable=tempNowF, font=("SF Pro Display", 28, "bold")).grid(
        row=1, column=0, sticky="w", padx=12, pady=(4, 0)
    )
    ctk.CTkLabel(tempCard, textvariable=tempSetF, font=("SF Pro Text", 14), text_color="gray40").grid(
        row=2, column=0, sticky="w", padx=12, pady=(2, 12)
    )

    # flow card
    flowCard = ctk.CTkFrame(content, corner_radius=16)
    flowCard.grid(row=0, column=1, sticky="nsew", padx=10, pady=10)
    ctk.CTkLabel(flowCard, text="Flow Rate (L/min)", font=("SF Pro Text",16,"bold")).grid(row=0, column=1, sticky="w", padx=12, pady=(12,4))
    flow = ctk.StringVar(value="—")
    ctk.CTkLabel(flowCard, textvariable=flow, font=("SF Pro Text", 20)).grid(row=1, column=0, sticky="w", padx=12, pady=(4, 12))

    # turbidity card
    turbCard = ctk.CTkFrame(content, corner_radius=16)
    turbCard.grid(row=1, column=0, sticky="nsew", padx=10, pady=10)
    ctk.CTkLabel(turbCard, text="Turbidity", font=("SF Pro Text",16,"bold")).grid(row=0, column=0, sticky="w", padx=12, pady=(12,4))
    turbPct = ctk.StringVar(value="—")
    ntu = ctk.StringVar(value="—")
    ctk.CTkLabel(turbCard, textvariable=turbPct, font=("SF Pro Text", 24)).grid(row=1, column=0, sticky="w", padx=12, pady=4)
    ctk.CTkLabel(turbCard, textvariable=ntu, font=("SF Pro Text", 14), text_color="gray40").grid(row=2, column=0, sticky="w", padx=12, pady=(0, 12))

    # level card
    levelCard = ctk.CTkFrame(content, corner_radius=16)
    levelCard.grid(row=1, column=1, sticky="nsew", padx=10, pady=10)
    ctk.CTkLabel(levelCard, text="Liquid Level", font=("SF Pro Text",16,"bold")).grid(row=0, column=0, sticky="w", padx=12, pady=(12,4))
    levelText = ctk.StringVar(value="—")
    ctk.CTkLabel(levelCard, textvariable=levelText, font=("SF Pro Text", 20)).grid(row=1, column=0, sticky="w", padx=12, pady=(4,12))

    # heater card
    heaterCard = ctk.CTkFrame(content, corner_radius=16)
    heaterCard.grid(row=2, column=0, columnspan=2, sticky="nsew", padx=10, pady=10)
    ctk.CTkLabel(heaterCard, text="Heater", font=("SF Pro Text",16,"bold")).grid(row=0, column=0, sticky="w", padx=12, pady=(12,6))

    heaterEnableVar = ctk.BooleanVar(value=False)
    heaterStatusText = ctk.StringVar(value="Disabled")
    heaterSetText = ctk.StringVar(value="Set: —")
    heaterOutText = ctk.StringVar(value="Output: —")
    heaterWarnText = ctk.StringVar(value="")
    lidText = ctk.StringVar(value="Lid: —")

    def on_heater_toggle():
        send(f"HEATER {1 if heaterEnableVar.get() else 0}")

    ctk.CTkSwitch(heaterCard, text="Enable", variable=heaterEnableVar, command=on_heater_toggle).grid(row=0, column=1, sticky="e", padx=12, pady=(12,6))
    ctk.CTkLabel(heaterCard, textvariable=lidText, font=("SF Pro Text", 14), text_color="gray40").grid(row=1, column=1, sticky="e", padx=12, pady=(0,4))
    ctk.CTkLabel(heaterCard, textvariable=heaterStatusText, font=("SF Pro Text", 14)).grid(row=1, column=0, sticky="w", padx=12, pady=(0,4))
    ctk.CTkLabel(heaterCard, textvariable=heaterSetText, font=("SF Pro Text", 14), text_color="gray40").grid(row=2, column=0, sticky="w", padx=12, pady=(0,4))
    ctk.CTkLabel(heaterCard, textvariable=heaterOutText, font=("SF Pro Text", 14), text_color="gray40").grid(row=3, column=0, sticky="w", padx=12, pady=(0,8))
    ctk.CTkLabel(heaterCard, textvariable=heaterWarnText, font=("SF Pro Text", 14, "bold"), text_color="red").grid(row=4, column=0, columnspan=2, sticky="w", padx=12, pady=(0,12))
    heaterCard.grid_columnconfigure(0, weight=1)

    # --------------------------
    # Queue pump
    # --------------------------
    def poll_queue():
        try:
            while True:
                kind, payload = q.get_nowait()
                if kind == "status":
                    title_status.set(payload)
                elif kind == "data":
                    try:
                        m = json.loads(payload)
                        if m.get("ok"):
                            tempNowF.set(f"{float(m.get('F', 0)):.1f} °F")
                            flow.set(f"{m.get('flowLpm', 0):.2f} L/min")
                            turbPct.set(f"{m.get('% Turbidity', 0):.1f} %")
                            ntu.set(f"{m.get('NTU', 0):.1f} NTU")

                            h_en   = m.get("heaterEnable", None)
                            h_set  = m.get("setTempF", None)
                            if h_set is not None:
                                tempSetF.set(f"Set: {float(h_set):.1f} °F")
                            else:
                                tempSetF.set("Set: —")
                            h_dem  = m.get("heaterDemand", None)
                            h_on   = m.get("heaterOn", None)
                            h_over = m.get("overTemp", None)
                            h_sec  = m.get("secOver", None)
                            h_stop = m.get("heaterStop", None)
                            lvl = m.get("levelOk", None)
                            lid_closed = m.get("lidClosed", None)

                            if isinstance(lid_closed, bool):
                                lidText.set("Lid is closed" if lid_closed else "Lid is open")
                                if not lid_closed and heaterEnableVar.get():
                                    heaterEnableVar.set(False)
                            else:
                                lidText.set("Lid status error!")

                            if lvl is None:
                                levelText.set("—")
                            elif lvl:
                                levelText.set("OK")
                            else:
                                levelText.set("LOW / FAULT")

                            if isinstance(h_en, bool):
                                heaterEnableVar.set(h_en)

                            if h_set is not None:
                                heaterSetText.set(f"Temp set to: {float(h_set):.1f} °F")
                            else:
                                heaterSetText.set("Temp set to: —")

                            status_parts = []
                            if isinstance(h_en, bool):
                                status_parts.append("Enabled" if h_en else "Disabled")
                            if isinstance(h_dem, bool):
                                status_parts.append(f"Demand: {'ON' if h_dem else 'OFF'}")
                            if isinstance(h_on, bool):
                                status_parts.append(f"Output: {'ON' if h_on else 'OFF'}")

                            heaterStatusText.set(" | ".join(status_parts) if status_parts else "—")
                            heaterOutText.set(f"Output: {'ON' if h_on else 'OFF'}" if isinstance(h_on, bool) else "Output: —")
                            timeLeft = 21 - h_sec if isinstance(h_sec, (int, float)) else None

                            if h_stop:
                                heaterWarnText.set("SHUTDOWN: Operating temp too high. Resume cycle once temp is below 120°F.")
                            elif h_over:
                                heaterWarnText.set(
                                    f"WARNING: Temp exceeded 120°F. System will shut down in {timeLeft} sec if system does not cool"
                                    if timeLeft is not None else
                                    "WARNING: ≥120°F"
                                )
                            else:
                                heaterWarnText.set("")

                            logger.log_event({
                                "timestamp": now_iso(),
                                "event_type": "telemetry",
                                "raw": "",

                                "C": m.get("C", ""),
                                "F": m.get("F", ""),
                                "flowLpm": m.get("flowLpm", ""),
                                "turbidity_pct": m.get("% Turbidity", ""),
                                "NTU": m.get("NTU", ""),

                                "heaterEnable": h_en,
                                "setTempF": h_set,
                                "heaterDemand": h_dem,
                                "heaterOn": h_on,
                                "overTemp": h_over,
                                "secOver": h_sec,
                                "heaterStop": h_stop,
                                "levelOk": lvl,
                                "lidClosed": lid_closed,

                                "ui_heater_enable": heaterEnableVar.get(),
                                "ui_cycle_state": cycle_state.get(),
                                "ui_cycle_remaining_s": cycle_remaining_s.get(),
                                "ui_cycle_duration_s": int(cycle_duration_min.get() * 60),
                                "ui_cycle_temp_set_f": float(cycle_temp_set_f.get()),
                            })

                    except json.JSONDecodeError:
                        title_status.set("Bad JSON")
                        log_status(f"Bad JSON: {payload[:200]}")
        except queue.Empty:
            pass
        app.after(200, poll_queue)

    threading.Thread(target=reader, daemon=True).start()
    app.after(200, poll_queue)

    app.after(250, comms_watchdog)

    show_page("operation")
    app.mainloop()

if __name__ == "__main__":
    main()
