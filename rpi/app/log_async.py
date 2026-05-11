"""Async logging utilities for telemetry and UI events."""
import os, csv, time, queue, threading
from typing import Any, Dict, List, Optional
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from app.state import AppModel


class AsyncLogger:
    """Append structured log rows to CSV and XLSX on a background thread."""
    def __init__(
        self,
        xlsx_path: str,
        csv_path: str,
        event_headers: List[str],
        max_dir_bytes: Optional[int] = None,
        prune_target_bytes: Optional[int] = None,
    ):
        """Initialize the logger and start the background worker."""
        self.xlsx_path = xlsx_path
        self.csv_path = csv_path
        self.event_headers = event_headers
        self.log_dir = os.path.dirname(self.csv_path) or "."
        self.max_dir_bytes = max_dir_bytes
        self.prune_target_bytes = prune_target_bytes if prune_target_bytes is not None else max_dir_bytes

        self.lq: "queue.Queue[dict]" = queue.Queue()
        self.stop_flag = threading.Event()
        self._sentinel = object()

        self._init_csv()
        self._init_xlsx()
        self._prune_if_needed()

        self.t = threading.Thread(target=self._worker, daemon=True)
        self.t.start()

    def _init_csv(self):
        """Create the CSV log file with headers if needed."""
        if not os.path.exists(self.csv_path):
            with open(self.csv_path, "w", newline="", encoding="utf-8") as f:
                csv.writer(f).writerow(self.event_headers)

    def _init_xlsx(self):
        """Create the XLSX log file and worksheets if needed."""
        if os.path.exists(self.xlsx_path):
            return

        wb = Workbook()
        ws_events = wb.active
        ws_events.title = "Events"
        ws_events.append(self.event_headers)

        ws_rx = wb.create_sheet("RxRaw")
        ws_rx.append(["timestamp", "raw_line"])

        ws_tx = wb.create_sheet("TxCmd")
        ws_tx.append(["timestamp", "command"])

        for ws in (ws_events, ws_rx, ws_tx):
            for i, h in enumerate(ws[1], start=1):
                ws.column_dimensions[get_column_letter(i)].width = max(12, min(42, len(str(h)) + 2))

        wb.save(self.xlsx_path)

    def log_event(self, row: dict):
        """Queue a single event row for persistence."""
        self.lq.put(row)

    # Convenience wrappers so main.py stays clean
    def log_status(self, timestamp: str, msg: str):
        """Record a status message in the event log."""
        self.log_event({"timestamp": timestamp, "event_type": "status", "raw": msg})

    def log_rx(self, timestamp: str, raw_line: str):
        """Record a raw RX line in the event log."""
        self.log_event({"timestamp": timestamp, "event_type": "rx_raw", "raw": raw_line})

    def log_tx(self, timestamp: str, cmd: str):
        """Record a TX command in the event log."""
        self.log_event({"timestamp": timestamp, "event_type": "tx_cmd", "raw": cmd})

    def _append_xlsx(self, batch_rows: List[dict]):
        """Append a batch of event rows to the XLSX file."""
        wb = load_workbook(self.xlsx_path)
        ws_events = wb["Events"]
        ws_rx = wb["RxRaw"]
        ws_tx = wb["TxCmd"]

        for r in batch_rows:
            et = r.get("event_type", "")
            ts = r.get("timestamp", "")
            raw = r.get("raw", "")

            ws_events.append([r.get(k, "") for k in self.event_headers])

            if et == "rx_raw":
                ws_rx.append([ts, raw])
            elif et == "tx_cmd":
                ws_tx.append([ts, raw])

        wb.save(self.xlsx_path)

    def _append_csv(self, batch_rows: List[dict]):
        """Append a batch of event rows to the CSV file."""
        with open(self.csv_path, "a", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            for r in batch_rows:
                w.writerow([r.get(k, "") for k in self.event_headers])

    def _prune_if_needed(self):
        """Delete the oldest inactive log files once the log directory grows too large."""
        if not self.max_dir_bytes or not os.path.isdir(self.log_dir):
            return

        keep_paths = {os.path.abspath(self.csv_path), os.path.abspath(self.xlsx_path)}
        log_files = []
        total_bytes = 0

        for name in os.listdir(self.log_dir):
            if not name.endswith((".csv", ".xlsx")):
                continue
            path = os.path.join(self.log_dir, name)
            if not os.path.isfile(path):
                continue
            try:
                size = os.path.getsize(path)
                mtime = os.path.getmtime(path)
            except OSError:
                continue
            total_bytes += size
            log_files.append((mtime, size, path))

        if total_bytes <= self.max_dir_bytes:
            return

        target_bytes = self.prune_target_bytes or self.max_dir_bytes
        for _, size, path in sorted(log_files):
            if total_bytes <= target_bytes:
                break
            if os.path.abspath(path) in keep_paths:
                continue
            try:
                os.remove(path)
                total_bytes -= size
            except OSError:
                continue

    def _worker(self):
        """Batch events and flush to disk on size or time thresholds."""
        batch: List[dict] = []
        last_flush = time.time()
        FLUSH_EVERY_SEC = 2.0
        FLUSH_EVERY_ROWS = 25

        while True:
            try:
                item = self.lq.get(timeout=0.25)
                if item is self._sentinel:
                    break
                batch.append(item)
            except queue.Empty:
                if self.stop_flag.is_set():
                    break

            do_flush = False
            if batch and len(batch) >= FLUSH_EVERY_ROWS:
                do_flush = True
            if batch and (time.time() - last_flush) >= FLUSH_EVERY_SEC:
                do_flush = True

            if do_flush:
                try:
                    self._append_csv(batch)
                    self._append_xlsx(batch)
                except Exception:
                    # fallback to CSV only
                    try:
                        self._append_csv(batch)
                    except Exception:
                        pass
                self._prune_if_needed()
                batch.clear()
                last_flush = time.time()

        if batch:
            try:
                self._append_csv(batch)
                self._append_xlsx(batch)
            except Exception:
                try:
                    self._append_csv(batch)
                except Exception:
                    pass
            self._prune_if_needed()

    def stop(self):
        """Signal the worker thread to stop."""
        self.stop_flag.set()
        self.lq.put(self._sentinel)
        self.t.join(timeout=3.0)

def now_iso():
    """Return a millisecond-precision timestamp string."""
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")[:-3]

def log_telemetry(logger: AsyncLogger, model: AppModel, ui: Dict[str, Any]) -> None:
    """Emit a telemetry event row from the current model state."""
    t = model.telemetry
    if not t.ok:
        return

    logger.log_event({
        "timestamp": ui.get("timestamp", ""),
        "event_type": "telemetry",
        "raw": "",

        "C": t.C,
        "F": t.F,
        "flowLpm": t.flowLpm,
        "heaterEnable": t.heaterEnable,
        "setTempF": t.setTempF,
        "heaterDemand": t.heaterDemand,
        "heaterOn": t.heaterOn,
        "overTemp": t.overTemp,
        "secOver": t.secOver,
        "heaterStop": t.heaterStop,
        "levelOk": t.levelOk,
        "lidClosed": t.lidClosed,

        "ui_cycle_state": ui.get("ui_cycle_state"),
        "ui_cycle_remaining_s": ui.get("ui_cycle_remaining_s"),
        "ui_cycle_duration_s": ui.get("ui_cycle_duration_s"),
        "ui_cycle_temp_set_f": ui.get("ui_cycle_temp_set_f"),
    })

def log_cycle_snapshot(logger: AsyncLogger, ui: Dict[str, Any]) -> None:
    """Record the current UI cycle configuration and state."""
    logger.log_event({
        "timestamp": now_iso(),
        "event_type": "status",
        "raw": "UI cycle state update",
        "ui_cycle_state": ui.cycle_state.get(),
        "ui_cycle_remaining_s": ui.cycle_remaining_s.get(),
        "ui_cycle_duration_s": int(ui.cycle_duration_min_in.get() * 60),
        "ui_cycle_temp_set_f": float(ui.cycle_temp_set_f_in.get()),
    })
